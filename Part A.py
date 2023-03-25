"""
Aircraft Collision Data Analysis, Part A

Description:
    This code processes aircraft collision data and creates a summary sheet with information on the
number of collisions with different animal species. It also includes a bar chart that displays the data. The code
removes unknown values from a specified column and splits the species name from a specified column, keeping only the
last word.

Author:
    Vadim Rudenko

Date:
    24 March 2023

Requirements:
    - Python 3
    - openpyxl library

Inputs:
    - smallAircraftData.xlsx: A Microsoft Excel workbook containing aircraft collision data

Outputs:
    - smallAircraftData.xlsx: Renewed file with a chart and filtered data

Methods:
    - remove_unknown_values(column_letter): Removes rows containing unknown
values from a specified column.
    - split_species_name(column_letter): Splits the species name from a
specified column and keeps only the last word.
    - create_animals_summary(): Creates a summary sheet containing information on the number of
collisions with different animal species and a filter for data.
    - save_changes(): Save changes made to the
Excel workbook.
    - print_most_collisions_species(): Prints the species with the most
collisions and the number of collisions.

"""

import openpyxl
from openpyxl.chart import BarChart, Reference


class AircraftData:
    def __init__(self, file):
        """
                Constructor for the AircraftData class.

                :param file: Name of the file containing the aircraft data.
        """
        self.collisions_dict = None
        self.most_collisions_species = None
        self.name_of_file = file
        self.workbook = openpyxl.load_workbook(self.name_of_file)
        self.sheet = self.workbook.active

    def remove_unknown_values(self, column_letter):
        """
               Removes rows containing unknown values from a specified column.

               :param column_letter: The letter representing the column to be processed.
        """
        rows_to_remove = []
        for cell in self.sheet[column_letter]:
            if cell.value is None or "UNKNOWN" in str(cell.value).upper():
                rows_to_remove.append(cell.row)
        for row in reversed(rows_to_remove):
            self.sheet.delete_rows(row)

    def split_species_name(self, column_letter):
        """
                Splits the species name from a specified column and keeps only the last word.

                :param column_letter: The letter representing the column to be processed.
        """
        for cell in self.sheet[column_letter]:
            if cell.value is not None:
                cell.value = cell.value.split()[-1]

    def create_animals_summary(self):
        """
        Creates a summary sheet containing information on the number of collisions with different animal species and
        filter data.
        """
        animals = [cell.value for cell in self.sheet["AF"] if cell.value is not None and cell.row != 1]

        if len(set(animals)) <= 15:
            animals_to_include = set(animals)
        else:
            collisions_dict = {animal: animals.count(animal) for animal in set(animals)}
            max_collisions = max(collisions_dict.values())
            animals_to_include = {animal for animal, count in collisions_dict.items() if count > 0.1 * max_collisions}

        most_collisions = max(animals_to_include, key=animals.count)
        sorted_collisions = sorted([(animal, animals.count(animal)) for animal in animals_to_include],
                                   key=lambda x: x[0], reverse=False)

        summary_sheet = self.workbook.create_sheet("ChartForAnimals")
        summary_sheet.title = "ChartForAnimals"
        summary_sheet["A1"] = "Name of Species"
        summary_sheet["B1"] = "Number of collisions"

        for i, (animal, count) in enumerate(sorted_collisions):
            summary_sheet.cell(row=i + 2, column=1).value = animal
            summary_sheet.cell(row=i + 2, column=2).value = count

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Collisions with animals"
        chart.y_axis.title = 'Number of collisions'
        chart.x_axis.title = 'Species'

        data_range = Reference(summary_sheet, min_col=2, min_row=1, max_row=len(sorted_collisions) + 1, max_col=3)
        categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=len(sorted_collisions) + 1)

        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4

        summary_sheet.add_chart(chart, "E1")

        self.most_collisions_species = most_collisions
        self.collisions_dict = {animal: animals.count(animal) for animal in animals_to_include}

    def save_changes(self):
        """
        Save changes made to the Excel workbook.
        :return: None
        """
        self.workbook.save(self.name_of_file)

    def print_most_collisions_species(self):
        """
           Print the species with the most collisions and the number of collisions.
           :return: None
        """
        print(
            f"The species with the most collisions is {self.most_collisions_species}: "
            f"{self.collisions_dict[self.most_collisions_species]}")


name_of_file = "smallAircraftData.xlsx"
aircraft_data = AircraftData(name_of_file)
aircraft_data.remove_unknown_values("AF")
aircraft_data.split_species_name("AF")
aircraft_data.create_animals_summary()
aircraft_data.save_changes()
aircraft_data.print_most_collisions_species()
