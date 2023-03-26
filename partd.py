"""
Aircraft Collision Data Analysis, Part D
Description:
    This script loads a Microsoft Excel workbook containing aircraft collision data and generates a bar chart
showing the frequency of collisions by month.
Author:
    Harsh Darji
Date:
    24 March 2023
Requirements:
    - Python 3
    - openpyxl library
Inputs:
    - aircraftWildlifeStrikes.xlsx: A Microsoft Excel workbook containing aircraft collision data
Outputs:
    - ChartForMonths worksheet: A new worksheet in the input workbook containing a bar chart showing the
frequency of collisions by month
Functions:
    - get_month_cells(cells: list) -> list: Extracts the values in the Incident Month column of the input
worksheet
    - get_cell_values(months: list) -> list: Extracts the cell coordinates in the Incident Month column of the
input worksheet
    - get_cells_and_months(incident_months: list, month_cells: list) -> list: Combines the cell
coordinates and values of the Incident Month column
    - get_count_of_months(list_cells_months: list) -> list: Counts
the frequency of each month in the Incident Month column
"""



from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

workbook = load_workbook("aircraftWildlifeStrikes.xlsx")
sheet2 = workbook.create_sheet(title="ChartForAirlines")

firstSheet = workbook[workbook.sheetnames[0]]


def get_airline_cells(cells: list) -> list:
   """
       Returns a list of all the values in the 6th column of the first sheet of the workbook.
       
       Args:
       cells (list): an empty list to which the values will be appended.
       Returns:
       
       list: a list of all the values in the 6th column of the first sheet of the workbook.
    """
    
    for column_values in firstSheet:
        cells.append(column_values[5].value)
    return cells


def get_cell_values(airlines: list) -> list:
    """
        Returns a list of all the coordinate values in the 6th column of the first sheet of the workbook.

        Args:
        months (list): an empty list to which the coordinate values will be appended.

        Returns:
        list: a list of all the coordinate values in the 6th column of the first sheet of the workbook.
    """
    for airline in firstSheet:
        airlines.append(airline[5].coordinate)
    return airlines


def get_cells_and_airlines(incident_airlines: list, airline_cells: list) -> list:
    """
       Returns a list of lists, where each list contains a coordinate value from the 6th column of the first sheet
    of the workbook and its corresponding value from the same row in the third column.

       Args:
       incident_months (list): an empty list to which the coordinate
    values from the second column will be appended.
       month_cells (list): an empty list to which the values from
    the third column will be appended.

       Returns:
       list: a list of lists, where each list contains a coordinate value from the 6th column of the first sheet
    of the workbook and its corresponding value from the same row in the third column.
    """
    return list(
        zip(get_cell_values(incident_airlines), get_airline_cells(airline_cells)))


def get_count_of_airlines(list_cells_airlines: list) -> list:
    """
        Returns a list of lists, where each sublist contains a unique value from the 6th column of the first sheet
    of the workbook and its count in the column.

        Args:
        list_cells_months (list): a list of lists, where each list contains a coordinate value from the 6th column
    of the first sheet of the workbook and its corresponding value from the same row in the third column.

        Returns:
        list: a list of lists, where each sublist contains a unique value from the second column of the first sheet
    of the workbook and its count in the column.
    """
    airlines = [list_cells_airlines[i][1]
                for i in range(len(list_cells_airlines))]

    return [[airline, airlines.count(airline)] for airline in set(airlines)]


airline_cells = []
incident_airlines = []
rows = get_count_of_airlines(
    get_cells_and_airlines(incident_airlines, airline_cells))
rows.remove(['Operator', 1])

# credit for code below - https://stackoverflow.com/questions/65679123/sort-nested-list-data-in-python
rows = sorted(rows, key=lambda x: x[0])
print(rows)
print("\n\n\n")
print(len(rows))
print("\n\n\n")

max_collisions = max(rows, key=lambda x: x[1]) # getting maximum number of collisions
print(max_collisions)
# print(type(max_collisions))

len_of_rows = len(rows)

row_no = 0

while row_no < len_of_rows:
    if 'UNKNOWN' in rows[row_no]: # deleting unknown airlines
        rows.pop(row_no)
        len_of_rows -= 1
        row_no -= 1
    elif rows[row_no][1] < (max_collisions[1]*0.1): # deleting flights with number of number of collisions less than 10% of maximum
        rows.pop(row_no)
        len_of_rows -= 1
        row_no -= 1
    row_no += 1


# print(rows)


for row in rows:
    sheet2.append(row)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Airlines and # of Collisions"
chart1.y_axis.title = 'Frequency'
chart1.x_axis.title = 'Airlines'

data = Reference(sheet2, min_col=2, min_row=1, max_row=len(rows), max_col=2)
cats = Reference(sheet2, min_col=1, min_row=1, max_row=len(rows), max_col=1)
chart1.add_data(data)
chart1.set_categories(cats)
chart1.legend = None
sheet2.add_chart(chart1, "D2")
workbook.save('aircraftWildlifeStrikes.xlsx') # save data
