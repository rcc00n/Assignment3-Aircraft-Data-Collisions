"""
Aircraft Collision Data Analysis, Part C

Description:
    This script loads a Microsoft Excel workbook containing aircraft collision data and generates a bar chart
showing the frequency of collisions by month.

Author:
    Harsh Darji

Date:
    24 March 2023

Requirements:
    - Python 3
    - openpyxl library

Inputs:
    - aircraftWildlifeStrikes.xlsx: A Microsoft Excel workbook containing aircraft collision data

Outputs:
    - ChartForMonths worksheet: A new worksheet in the input workbook containing a bar chart showing the
frequency of collisions by month

Functions:
    - get_month_cells(cells: list) -> list: Extracts the values in the Incident Month column of the input
worksheet
    - get_cell_values(months: list) -> list: Extracts the cell coordinates in the Incident Month column of the
input worksheet
    - get_cells_and_months(incident_months: list, month_cells: list) -> list: Combines the cell
coordinates and values of the Incident Month column
    - get_count_of_months(list_cells_months: list) -> list: Counts
the frequency of each month in the Incident Month column

"""

from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

workbook = load_workbook("aircraftWildlifeStrikes.xlsx")
sheet2 = workbook.create_sheet(title="ChartForMonths")

firstSheet = workbook[workbook.sheetnames[0]]


def get_month_cells(cells: list) -> list:
    """
       Returns a list of all the values in the 3rd column of the first sheet of the workbook.

       Args:
       cells (list): an empty list to which the values will be appended.

       Returns:
       list: a list of all the values in the second column of the first sheet of the workbook.
    """
    for column_values in firstSheet:
        cells.append(column_values[2].value)
    return cells


def get_cell_values(months: list) -> list:
    """
        Returns a list of all the coordinate values in the 3rd column of the first sheet of the workbook.

        Args:
        months (list): an empty list to which the coordinate values will be appended.

        Returns:
        list: a list of all the coordinate values in the second column of the first sheet of the workbook.
    """
    for month in firstSheet:
        months.append(month[2].coordinate)
    return months


def get_cells_and_months(incident_months: list, month_cells: list) -> list:
    """
       Returns a list of lists, where each list contains a coordinate value from the 3rd column of the first sheet
    of the workbook and its corresponding value from the same row in the third column.

       Args:
       incident_months (list): an empty list to which the coordinate
    values from the second column will be appended.
       month_cells (list): an empty list to which the values from
    the third column will be appended.

       Returns:
       list: a list of lists, where each list contains a coordinate value from the second column of the first sheet
    of the workbook and its corresponding value from the same row in the third column.
    """

    return list(
        zip(get_cell_values(incident_months), get_month_cells(month_cells)))


def get_count_of_months(list_cells_months: list) -> list:
    """
        Returns a list of lists, where each sublist contains a unique value from the second column of the first sheet
    of the workbook and its count in the column.

        Args:
        list_cells_months (list): a list of tuples, where each tuple contains a coordinate value from the second column
    of the first sheet of the workbook and its corresponding value from the same row in the third column.

        Returns:
        list: a list of lists, where each sublist contains a unique value from the second column of the first sheet
    of the workbook and its count in the column.
    """
    months = [list_cells_months[i][1] for i in range(len(list_cells_months))]

    # return set(months)
    return [[month, months.count(month)] for month in set(months)]


month_cells = []
incident_months = []
rows = get_count_of_months(get_cells_and_months(incident_months, month_cells))
rows.remove(['Incident Month', 1])

for row in rows:
    sheet2.append(row)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Months and # of Collisions"
chart1.y_axis.title = 'Frequency'
chart1.x_axis.title = 'Months'

data = Reference(sheet2, min_col=2, min_row=1, max_row=len(rows), max_col=2)
cats = Reference(sheet2, min_col=1, min_row=1, max_row=len(rows), max_col=1)
chart1.add_data(data)
chart1.set_categories(cats)
chart1.legend = None
sheet2.add_chart(chart1, "C1")
workbook.save("aircraftWildlifeStrikes.xlsx")
