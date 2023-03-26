"""
Aircraft Collision Data Analysis, Part B

Description:
    This script reads data from an Excel file containing information about small aircraft collisions
and creates a new sheet with a bar chart showing the frequency of collisions per year.

Author:
    Harsh Darji

Date:
    24 March 2023

Requirements:
    - Python 3
    - openpyxl library

Inputs:
    - smallAircraftData.xlsx: A Microsoft Excel workbook containing aircraft collision data

Outputs:
    - ChartForYears worksheet: A sheet in the input file containing a bar chart
showing the frequency of collisions per year.

Functions:
    - get_year_cells(cells: list) -> list: Returns a list of years from the first
column of the first sheet.
    - get_cell_values(years: list) -> list: Returns a list of cell coordinates in
the second column of the first sheet.
    - get_cells_and_years(incident_years: list, year_cells: list) -> list: Returns a list of tuples
containing cell coordinates and corresponding years.
    - get_count_of_years(list_cells_years: list) -> list: Returns a list of lists, each containing a
year and the count of its occurrences in the list.

"""

from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart

workbook = load_workbook("aircraftWildlifeStrikes.xlsx")
sheet2 = workbook.create_sheet(title="ChartForYears")

firstSheet = workbook[workbook.sheetnames[0]]


def get_year_cells(cells: list) -> list:
    """
        Returns a list of years from the 2nd column of the first sheet.

        Args:
            cells (list): An empty list to store the years.

        Returns:
            A list of years from the 2nd column of the first sheet.
    """
    for column_values in firstSheet:
        cells.append(column_values[1].value)
    return cells


def get_cell_values(years: list) -> list:
    """
        Returns a list of cell coordinates in the second column of the first sheet.

        Args:
            years (list): An empty list to store the cell coordinates.

        Returns:
            A list of cell coordinates in the second column of the first sheet.
    """
    for year in firstSheet:
        years.append(year[1].coordinate)
    return years


def get_cells_and_years(incident_years: list, year_cells: list) -> list:
    """
        Returns a list of lists containing cell coordinates and corresponding years.

        Args:
            incident_years (list): An empty list to store the incident years.
            year_cells (list): An empty list to store the year cell coordinates.

        Returns:
            A list of lists containing cell coordinates and corresponding years.
    """
    return list(
        zip(get_cell_values(incident_years), get_year_cells(year_cells)))


def get_count_of_years(list_cells_years: list) -> list:
    """
       Returns a list of lists, each containing a year and the count of its occurrences in the list.

       Args:
           list_cells_years (list): A list of tuples containing cell coordinates and corresponding years.

       Returns:
           A list of lists, each containing a year and the count of its occurrences in the list.
    """
    years = [list_cells_years[i][1] for i in range(len(list_cells_years))]

    # return set(years)
    return [[year, years.count(year)] for year in set(years)]


year_cells = []
incident_years = []
rows = get_count_of_years(get_cells_and_years(incident_years, year_cells))
rows.remove(['Incident Year', 1])

for row in rows:
    sheet2.append(row)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Years and # of Collisions"
chart1.y_axis.title = 'Frequency'
chart1.x_axis.title = 'Years'

data = Reference(sheet2, min_col=2, min_row=1, max_row=len(rows), max_col=2)
cats = Reference(sheet2, min_col=1, min_row=1, max_row=len(rows), max_col=1)
chart1.add_data(data)
chart1.set_categories(cats)
chart1.legend = None
sheet2.add_chart(chart1, "C1")
workbook.save("aircraftWildlifeStrikes.xlsx")
